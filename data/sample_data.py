import openpyxl

# 후킹성 블로그 포스팅 제목 10개
hook_titles = [
  '하루 5분, 인생이 바뀌는 아침 루틴 공개!',
  '이 방법 몰랐다면 손해! 집에서 돈 버는 꿀팁 10가지',
  '전문가가 알려주는 다이어트 성공 비법',
  '퇴근 후 1시간, 부수입 만드는 현실적인 방법',
  '누구나 쉽게 따라하는 스마트폰 사진 잘 찍는 법',
  '월급쟁이도 부자되는 투자 습관 7가지',
  '집안일이 2배 빨라지는 정리정돈 꿀팁',
  '초보도 성공하는 블로그 수익화 전략',
  '알아두면 쓸모있는 무료 온라인 강의 TOP 5',
  '직장인 필수! 시간 관리의 모든 것'
]

def main():
  try:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    # 헤더 입력
    ws['A1'] = '제목'
    ws['B1'] = '내용'
    # 제목 데이터 입력
    for idx, title in enumerate(hook_titles, start=2):
      ws[f'A{idx}'] = title
      ws[f'B{idx}'] = ''
    wb.save('/docs/data.xlsx')
    print('data.xlsx 파일이 성공적으로 생성되었습니다.')
  except Exception as e:
    print('에러 발생:', e)

if __name__ == '__main__':
  main() 