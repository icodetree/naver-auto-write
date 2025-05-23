import time
import openpyxl
import pyperclip
import random
import pyautogui
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.action_chains import ActionChains

# 네이버 계정 정보
NAVER_ID = 'master8407'
NAVER_PW = 'qusghdud0311#$'

# 네이버 로그인 함수
def naver_login(driver, wait):
  # 네이버 로그인 페이지 접속
  driver.get('https://nid.naver.com/nidlogin.login')
  
  # 영문 입력기로 전환 (Mac: Command + Space, Windows: Alt + Shift)
  if 'darwin' in __import__('platform').system().lower():  # Mac
    pyautogui.hotkey('command', 'space')
  else:  # Windows
    pyautogui.hotkey('alt', 'shift')
  time.sleep(0.2)
  
  # 아이디 입력 (클립보드 복사 방식)
  id_input = wait.until(EC.presence_of_element_located((By.ID, 'id')))
  id_input.click()
  time.sleep(0.5)
  pyperclip.copy(NAVER_ID)
  if 'darwin' in __import__('platform').system().lower():  # Mac
    pyautogui.hotkey('command', 'v')
  else:  # Windows
    pyautogui.hotkey('ctrl', 'v')
  time.sleep(0.5)
  
  # 비밀번호 입력 (클립보드 복사 방식)
  pw_input = wait.until(EC.presence_of_element_located((By.ID, 'pw')))
  pw_input.click()
  time.sleep(0.5)
  pyperclip.copy(NAVER_PW)
  if 'darwin' in __import__('platform').system().lower():  # Mac
    pyautogui.hotkey('command', 'v')
  else:  # Windows
    pyautogui.hotkey('ctrl', 'v')
  time.sleep(0.5)
  
  # 로그인 버튼 클릭
  login_btn = wait.until(EC.element_to_be_clickable((By.ID, 'log.login')))
  login_btn.click()
  time.sleep(3)  # 로그인 완료 대기
  
  # 새로운 브라우저 등록 팝업 처리
  try:
    # "다음에 하기" 버튼 찾기 시도 (버튼 텍스트가 다를 수 있음)
    skip_buttons = driver.find_elements(By.XPATH, "//button[contains(text(), '다음에 하기')]")
    if not skip_buttons:
      skip_buttons = driver.find_elements(By.XPATH, "//button[contains(text(), '취소')]")
    if not skip_buttons:
      skip_buttons = driver.find_elements(By.XPATH, "//button[contains(text(), '나중에')]")
    if not skip_buttons:
      # 클래스명으로 시도
      skip_buttons = driver.find_elements(By.CSS_SELECTOR, ".btn_next")
      
    if skip_buttons:
      skip_buttons[0].click()
      print("새 브라우저 등록 팝업 닫기 완료")
      time.sleep(1)
  except Exception as e:
    print("새 브라우저 팝업 처리 중 예외 발생 (무시됨):", e)
  
  print("네이버 로그인 완료")
  return True

# 글쓰기 페이지 준비 및 초기화 함수 (팝업/도움말 처리)
def prepare_editor(driver, wait):
  print("\n--- 글쓰기 페이지 준비 시작 ---")
  
  # 1. 글쓰기 페이지로 이동
  driver.get('https://blog.naver.com/GoBlogWrite.naver')
  time.sleep(3)  # 충분한 로딩 시간 확보
  print("글쓰기 페이지로 이동 완료")
  
  # 페이지 로드 확인
  WebDriverWait(driver, 10).until(lambda d: d.execute_script('return document.readyState') == 'complete')
  print("페이지 로드 완료")
  
  # 2. iframe 전환
  try:
    driver.switch_to.default_content()  # 먼저 최상위 프레임으로 전환
    
    # iframe ID로 직접 전환
    main_frame = wait.until(EC.presence_of_element_located((By.ID, 'mainFrame')))
    driver.switch_to.frame(main_frame)
    print("mainFrame iframe 전환 성공")
    time.sleep(2)
  except Exception as e:
    print(f"iframe 전환 실패 (계속 진행): {e}")
  
  # 3. 도움말 패널 닫기 (여러 방법 시도)
  close_help_panel(driver, wait)
  
  # 4. 다른 팝업 닫기
  try:
    # 작성 중인 글 팝업 있으면 닫기
    continue_buttons = driver.find_elements(By.XPATH, "//button[contains(text(), '새로 작성')]")
    if continue_buttons:
      for btn in continue_buttons:
        if btn.is_displayed():
          btn.click()
          print("새로 작성 버튼 클릭")
          time.sleep(1)
    
    # 취소 버튼 시도
    cancel_buttons = driver.find_elements(By.XPATH, "//button[contains(text(), '취소')]")
    for btn in cancel_buttons:
      if btn.is_displayed():
        btn.click()
        print("취소 버튼 클릭")
        time.sleep(1)
    
    # 시작하기/닫기 버튼 시도
    close_buttons = driver.find_elements(By.XPATH, 
      "//button[contains(text(), '시작하기') or contains(text(), '닫기') or contains(@class, 'close') or contains(@class, 'cancel')]"
    )
    for btn in close_buttons:
      if btn.is_displayed():
        btn.click()
        print("기타 버튼 클릭")
        time.sleep(1)
  except Exception as e:
    print(f"팝업 처리 중 오류 (계속 진행): {e}")
  
  # 5. 에디터 준비 상태 확인
  try:
    # 제목 필드 찾기 시도 (JavaScript)
    has_title_field = driver.execute_script("""
      return Boolean(
        document.querySelector('.se-section-documentTitle') || 
        document.querySelector('.se-documentTitle-input') ||
        document.querySelector('.document_title') ||
        document.querySelector('input.se-ff-nanummyeongjo')
      );
    """)
    
    if has_title_field:
      print("에디터 준비 완료: 제목 필드 확인됨")
      return True
    else:
      print("제목 필드를 찾을 수 없음, 다시 도움말 닫기 시도")
      close_help_panel(driver, wait)
      
      # 한 번 더 확인
      has_title_field = driver.execute_script("""
        return Boolean(
          document.querySelector('.se-section-documentTitle') || 
          document.querySelector('.se-documentTitle-input') ||
          document.querySelector('.document_title') ||
          document.querySelector('input.se-ff-nanummyeongjo')
        );
      """)
      
      if has_title_field:
        print("에디터 준비 완료: 제목 필드 확인됨 (2차 시도)")
        return True
      else:
        print("제목 필드를 여전히 찾을 수 없음")
        return False
  except Exception as e:
    print(f"에디터 준비 상태 확인 중 오류: {e}")
    return False

# 도움말 패널 닫기 함수
def close_help_panel(driver, wait):
  print("도움말 패널 닫기 시도...")
  
  # 1. JavaScript로 도움말 패널 확인 및 닫기
  driver.execute_script("""
    var helpCloseBtn = document.querySelector('.se-help-panel-close-button');
    if (helpCloseBtn) {
      console.log('도움말 패널 닫기 버튼 찾음');
      helpCloseBtn.click();
      return true;
    }
    return false;
  """)
  time.sleep(1)
  
  # 2. 직접 셀렉터로 찾아서 닫기
  try:
    help_buttons = driver.find_elements(By.CSS_SELECTOR, '.se-help-panel-close-button')
    if help_buttons:
      for btn in help_buttons:
        if btn.is_displayed():
          driver.execute_script("arguments[0].click();", btn)
          print("도움말 패널 닫기 버튼 클릭 성공")
          time.sleep(1)
  except:
    pass
  
  # 3. XPath로 시도
  try:
    xpath_buttons = driver.find_elements(By.XPATH, "//button[contains(@class, 'se-help-panel-close-button')]")
    if xpath_buttons:
      for btn in xpath_buttons:
        if btn.is_displayed():
          driver.execute_script("arguments[0].click();", btn)
          print("XPath로 도움말 패널 닫기 버튼 클릭")
          time.sleep(1)
  except:
    pass
  
  # 4. ESC 키 입력으로 닫기 시도
  try:
    actions = ActionChains(driver)
    actions.send_keys(Keys.ESCAPE).perform()
    print("ESC 키로 도움말 패널 닫기 시도")
    time.sleep(1)
  except:
    pass
    
  print("도움말 패널 닫기 완료")

# 글 작성 함수 (제목, 내용만 입력하고 저장)
def write_post(driver, wait, title, content):
  try:
    # ActionChains 초기화
    actions = ActionChains(driver)
    
    # 1. 제목 입력
    try:
      # 다양한 제목 필드 셀렉터 시도
      title_selectors = [
        '.se-section-documentTitle',
        '.se-documentTitle-input',
        '.document_title',
        'input.se-ff-nanummyeongjo'
      ]
      
      title_elem = None
      for selector in title_selectors:
        elems = driver.find_elements(By.CSS_SELECTOR, selector)
        for elem in elems:
          if elem.is_displayed():
            title_elem = elem
            break
        if title_elem:
          break
      
      if not title_elem:
        # XPath로 시도
        xpath_elems = driver.find_elements(By.XPATH, "//*[contains(@class, 'title') or contains(@placeholder, '제목')]")
        for elem in xpath_elems:
          if elem.is_displayed():
            title_elem = elem
            break
      
      if title_elem:
        title_elem.click()
        time.sleep(0.5)
        
        # 기존 내용 삭제
        actions.key_down(Keys.COMMAND if 'darwin' in __import__('platform').system().lower() else Keys.CONTROL)
        actions.send_keys('a')
        actions.key_up(Keys.COMMAND if 'darwin' in __import__('platform').system().lower() else Keys.CONTROL)
        actions.send_keys(Keys.DELETE)
        actions.perform()
        time.sleep(0.5)
        
        # 새 제목 입력
        print(f"제목 입력: {title}")
        for ch in title:
          actions.send_keys(ch).perform()
          time.sleep(random.uniform(0.01, 0.10))
        print("제목 입력 완료")
      else:
        raise Exception("제목 입력 필드를 찾을 수 없습니다")
    except Exception as e:
      print(f"제목 입력 중 오류: {e}")
      raise
    
    # 2. 본문 입력
    try:
      # 다양한 본문 필드 셀렉터 시도
      body_selectors = [
        '.se-section-text',
        '.se-text-paragraph',
        '.content_text',
        '.se-main-container'
      ]
      
      body_elem = None
      for selector in body_selectors:
        elems = driver.find_elements(By.CSS_SELECTOR, selector)
        for elem in elems:
          if elem.is_displayed():
            body_elem = elem
            break
        if body_elem:
          break
      
      if not body_elem:
        # XPath로 시도
        xpath_elems = driver.find_elements(By.XPATH, "//*[contains(@class, 'content') or contains(@class, 'text') or contains(@class, 'body')]")
        for elem in xpath_elems:
          if elem.is_displayed() and elem.get_attribute('contenteditable') == 'true':
            body_elem = elem
            break
      
      if body_elem:
        body_elem.click()
        time.sleep(0.5)
        
        # 기존 내용 삭제
        actions.key_down(Keys.COMMAND if 'darwin' in __import__('platform').system().lower() else Keys.CONTROL)
        actions.send_keys('a')
        actions.key_up(Keys.COMMAND if 'darwin' in __import__('platform').system().lower() else Keys.CONTROL)
        actions.send_keys(Keys.DELETE)
        actions.perform()
        time.sleep(0.5)
        
        # 새 본문 입력
        print("본문 입력 시작")
        # 본문 내용을 줄 단위로 분리하여 입력
        lines = content.split('\n')
        for line in lines:
          for ch in line:
            actions.send_keys(ch).perform()
            time.sleep(random.uniform(0.01, 0.10))
          actions.send_keys('\n').perform()
          time.sleep(0.1)
        print("본문 입력 완료")
      else:
        raise Exception("본문 입력 필드를 찾을 수 없습니다")
    except Exception as e:
      print(f"본문 입력 중 오류: {e}")
      raise
    
    # 3. 저장 버튼 클릭
    try:
      # 다양한 저장 버튼 셀렉터 시도
      save_selectors = [
        '.save_btn__bzc5B',
        '.save_btn',
        '.btn_publish',
        '.btn_save',
        '.se_publish',
        '.se-publish-button',
        '.publish_button'
      ]
      
      save_btn = None
      for selector in save_selectors:
        elems = driver.find_elements(By.CSS_SELECTOR, selector)
        for elem in elems:
          if elem.is_displayed():
            save_btn = elem
            break
        if save_btn:
          break
      
      if not save_btn:
        # XPath로 시도
        xpath_btns = driver.find_elements(By.XPATH, "//button[contains(text(), '발행') or contains(text(), '저장') or contains(text(), '등록')]")
        for btn in xpath_btns:
          if btn.is_displayed():
            save_btn = btn
            break
      
      if save_btn:
        print("저장 버튼 클릭")
        save_btn.click()
        
        # 저장 완료 확인 (팝업 또는 메시지 대기)
        try:
          wait.until(lambda driver: any(
            len(driver.find_elements(By.XPATH, f"//*[contains(text(), '{text}')]")) > 0
            for text in ['저장되었습니다', '발행되었습니다', '완료', '성공']
          ))
          print("저장 완료 확인됨")
        except:
          print("저장 완료 메시지를 확인할 수 없습니다. 5초 대기 후 계속 진행합니다.")
          time.sleep(5)
      else:
        raise Exception("저장 버튼을 찾을 수 없습니다")
    except Exception as e:
      print(f"저장 중 오류: {e}")
      raise
    
    # 다음 글 작성 준비
    driver.switch_to.default_content()
    time.sleep(3)  # 저장 후 충분한 대기 시간
    
    # 글쓰기 페이지로 다시 이동
    driver.get('https://blog.naver.com/GoBlogWrite.naver')
    time.sleep(3)
    
    # iframe 다시 전환
    try:
      driver.switch_to.default_content()
      main_frame = wait.until(EC.presence_of_element_located((By.ID, 'mainFrame')))
      driver.switch_to.frame(main_frame)
      time.sleep(2)
      
      # 도움말 패널 다시 닫기
      close_help_panel(driver, wait)
    except Exception as e:
      print(f"다음 글 작성 준비 중 오류: {e}")
    
    return True
  except Exception as e:
    print(f"글 작성 실패: {e}")
    return False

def main():
  # 엑셀 파일 열기
  wb = openpyxl.load_workbook('docs/data.xlsx')
  ws = wb.active
  max_row = ws.max_row
  print(f"엑셀 파일 열기 성공: {max_row}행 데이터 발견")
  
  # 셀레니움 옵션 설정
  options = webdriver.ChromeOptions()
  options.add_argument('--disable-blink-features=AutomationControlled')
  options.add_argument('--start-maximized')
  
  # 드라이버 실행 (자동 경로)
  service = Service(ChromeDriverManager().install())
  driver = webdriver.Chrome(service=service, options=options)
  wait = WebDriverWait(driver, 20)  # 타임아웃 시간 증가 (10→20초)
  
  try:
    # 1. 네이버 로그인
    if not naver_login(driver, wait):
      raise Exception("로그인 실패")
    
    # 2. 글쓰기 페이지 준비 (팝업, 도움말 등 초기 처리)
    if not prepare_editor(driver, wait):
      raise Exception("글쓰기 페이지 준비 실패")
    
    # 3. 엑셀 데이터로 글 작성
    success_count = 0
    content_generated_count = 0
    
    for idx in range(2, max_row + 1):
      title = ws[f'A{idx}'].value
      content = ws[f'B{idx}'].value
      
      # 제목이 없으면 건너뛰기
      if not title:
        print(f"행 {idx}: 제목이 없습니다. 건너뜁니다.")
        continue
      
      # 내용이 없으면 내용 자동 생성
      if not content:
        print(f"행 {idx}: 내용이 없습니다. 자동으로 내용을 생성합니다.")
        content = f"""[서론]
{title}에 대한 블로그 글입니다.
이 글은 자동으로 생성된 내용입니다.
블로그 글을 시작합니다.

[본론]
{title}에 대해 자세히 설명합니다.
이런 주제는 많은 사람들에게 유용한 정보가 될 수 있습니다.
블로그를 통해 정보를 공유하는 것은 매우 중요합니다.
다양한 의견과 경험을 나누어 볼 수 있습니다.
여러분의 의견도 댓글로 남겨주세요.

[결론]
오늘은 {title}에 대해 알아보았습니다.
앞으로도 유익한 정보로 찾아뵙겠습니다.
감사합니다.
"""
        print(f"내용 생성 완료: {len(content)}자")
        ws[f'B{idx}'] = content
        content_generated_count += 1
      
      print(f'\n--- {idx-1}/{max_row-1} 글 작성 시작: {title} ---')
      if write_post(driver, wait, title, content):
        print(f'행 {idx} 작성 완료: {title}')
        success_count += 1
        # 다음 글 작성 전 잠시 대기
        time.sleep(random.uniform(3, 5))
      else:
        print(f'행 {idx} 작성 실패: {title}')
        # 실패 시 글쓰기 페이지로 다시 이동하고 초기화
        try:
          driver.get('https://blog.naver.com/GoBlogWrite.naver')
          time.sleep(3)
          prepare_editor(driver, wait)
        except:
          print("글쓰기 페이지 복구 실패")
    
    # 엑셀 파일 저장 (내용 생성한 경우)
    if content_generated_count > 0:
      wb.save('docs/data.xlsx')
      print(f"\n자동 생성된 내용 {content_generated_count}개를 엑셀에 저장했습니다.")
    
    print(f'\n=== 작업 완료! 총 {max_row-1}개 중 {success_count}개 글이 성공적으로 작성되었습니다. ===')
    
  except Exception as e:
    print('에러 발생:', e)
  finally:
    input('엔터를 누르면 브라우저가 종료됩니다...')
    driver.quit()

if __name__ == '__main__':
  main() 