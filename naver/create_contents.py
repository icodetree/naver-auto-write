import openpyxl
import google.generativeai as genai

API_KEY = '구글 gemini api key'
genai.configure(api_key=API_KEY)
gemini_model = 'gemini-2.0-flash'

def generateBlogContent(title):
  prompt = f"""
  블로그 제목: {title}
  아래 형식에 맞춰 블로그 글을 작성해줘.
  - 서론: 3~4문장
  - 본론: 5~7문장
  - 결론: 2~3문장
  각 부분은 [서론], [본론], [결론]으로 구분해서 작성해줘.
  """
  try:
    model = genai.GenerativeModel(gemini_model)
    response = model.generate_content(prompt)
    return response.text.strip()
  except Exception as e:
    return f'생성 실패: {e}'

def main():
  try:
    wb = openpyxl.load_workbook('./docs/data.xlsx')
    ws = wb.active
    max_row = ws.max_row
    for idx in range(2, max_row + 1):
      title = ws[f'A{idx}'].value
      if not title:
        continue
      print(f'{idx-1}/{max_row-1} 처리 중: {title}')
      content = generateBlogContent(title)
      ws[f'B{idx}'] = content
    wb.save('./docs/data.xlsx')
    print('모든 블로그 콘텐츠 생성 및 저장 완료!')
  except Exception as e:
    print('에러 발생:', e)

if __name__ == '__main__':
  main() 